import openpyxl
import csv
import pandas as pd
import urllib.request
import time
import random
import re
from keras.preprocessing.text import Tokenizer
from sklearn.model_selection import train_test_split
import numpy as np
import json
import codecs









def write_one_row(data_list, filename):
    """
    write_one_row(['hash', 'vin_num', 'vout_num',  'in_address', 'out_address', 'type', 'value', 'fees', 'opreturn'], PATH_FAIL)
    :param data_list:
    :param filename:
    :return:
    """
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet1']
    treedata1 = [data_list]
    for row in range(len(treedata1)):
        ws.append(treedata1[row])
    wb.save(filename)
    print(f'FILE OUT.xlsx===============================================================has insert {data_list}')

def Z_ScoreNormalization(x,mu,sigma):
    x = (x - mu) / sigma;
    return x;

def get_data():
  
    filename_2 = "train.xlsx"
    filename_4 = "test.xlsx"

    normal_data = []
    special_data = []
    json_num = 0
    #处理json文件

    special_rawdata = "special_data.json"
    with open(special_rawdata, 'r', encoding='utf8')as fn:
        for line in fn.readlines():
            special = json.dumps(line)
            special = special.replace('\\n', '')
            special = special.replace('\\', '')
            special = special.replace('\"', '')
            special = special.replace('{', '')
            special = special.replace('}', '')
            special = special.replace('[', '')
            special = special.replace(']', '')
            special = special.replace(' ', '')
            special = special.replace(':', '')
            special = " ".join(special)
            special_data.append(special)

    normal_rawdata = "normal_data.json"
    with open(normal_rawdata, 'r', encoding='utf8')as fn:
        for line in fn.readlines():
            normal = json.dumps(line)
            normal = normal.replace('\\n', '')
            normal = normal.replace('\\', '')
            normal = normal.replace('\"', '')
            normal = normal.replace('{', '')
            normal = normal.replace('}', '')
            normal = normal.replace('[', '')
            normal = normal.replace(']', '')
            normal = normal.replace(' ', '')
            # normal = normal.replace('outputs', '')
            normal = normal.replace(':', '')
            normal = " ".join(normal)
            normal_data.append(normal)

   





    length_normal = len(normal_data)
    length_special= len(special_data)

   


    # ##CNN的训练和测试数据


    train_2 = openpyxl.load_workbook(filename_2)
    train_rawdata = train_2.active
    test_2 = openpyxl.load_workbook(filename_4)
    test_rawdata = test_2.active

    no = 0
    sp = 0
    num = 1 #记录表格的行数
    count = 0
    length = length_normal + length_special
    #
    #训练数据
    while count < (length * 0.7):



        print("train_normal")
        i = random.randint(1, 2)
        while i > 0 and no < (length_normal *0.7):
            string = normal_data[no]
            # string_1 = " ".join(string)
            train_rawdata.cell(row=num, column=1, value=string)
            train_rawdata.cell(row=num, column=2, value=int(0))
            # write_one_row(no_feat[no], filename_1)

            num = num + 1
            i = i - 1
            no = no + 1
            count = count + 1
            print(num)

        print("trian_special")
        j = random.randint(1, 2)
        while j > 0 and sp < (length_special *0.7):

            # train_rawdata.cell(row=num, column=1, value=gen_data[sp])
            # train_rawdata.cell(row=num, column=2, value=int(1))
            train_rawdata.cell(row=num, column=1, value=special_data[sp])
            train_rawdata.cell(row=num, column=2, value=int(1))

            num = num + 1
            j = j - 1
            sp = sp + 1
            count = count + 1
            print(num)

    # #
    # #
    # #
    # # #测试数据
    num = 1
    count = 0
    while count <(length * 0.3):
        print("test_normal")
        i = random.randint(1, 2)
        while i > 0 and no < length_normal:
            string = normal_data[no]
            # string_1 = " ".join(string)
            test_rawdata.cell(row=num, column=1, value=string)
            test_rawdata.cell(row=num, column=2, value=int(0))
            # write_one_row(no_feat[no], filename_3)

            num = num + 1
            i = i - 1
            no = no + 1
            count = count + 1
        print("trian_special")
        j = random.randint(1, 2)
        while j > 0 and sp < length_special:
            # test_rawdata.cell(row=num, column=1, value=gen_data[sp])
            # test_rawdata.cell(row=num, column=2, value=int(1))
            test_rawdata.cell(row=num, column=1, value=special_data[sp])
            test_rawdata.cell(row=num, column=2, value=int(1))
            num = num + 1
            j = j - 1
            sp = sp + 1
            count = count + 1
            print(num)






    train_2.save(filename_2)
    test_2.save(filename_4)






def load_CNN_data():

    #修改成放测试和训练数据的文档路径
    filename_2 = 'train.xlsx'

    filename_4 = 'test.xlsx'
    train_data = []
    train_label = []
    test_data = []
    test_label = []
    train = openpyxl.load_workbook(filename_2)
    train = train.active
    test = openpyxl.load_workbook(filename_4)
    test = test.active
    for cell in list(train.columns)[0]:  # 获取数据
        train_data.append(cell.value)

    for cell in list(train.columns)[1]:  # 获取标签
        train_label.append(cell.value)

    for cell in list(test.columns)[0]:  # 获取数据
        test_data.append(cell.value)
    for cell in list(test.columns)[1]:  # 获取标签
        test_label.append(cell.value)

    length_train = len(train_data)
    length_test = len(test_data)
    data = train_data
    data.extend(test_data)
    print(train_label)
    # tokenizer = Tokenizer(num_words=1000,filters='!"#$%&()*+,-./:;<=>?@[\]^_`{|}~\t\n',char_level=True)  # i创建一个分词器（tokenizer），设置为只考虑前1000个最常见的单词
    tokenizer = Tokenizer(num_words=1000)
    tokenizer.fit_on_texts(data)  # 构建索引单词

    sequences = tokenizer.texts_to_sequences(data)  # 将字符串转换为整数索引组成的列表
    data_all = []
    for string in sequences:
        data_all.append(string)

    train = data_all[0:length_train]
    test = data_all[length_train:(length_train+length_test)]

    return (train, train_label, test,  test_label)

