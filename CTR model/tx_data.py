import openpyxl
import csv
import pandas as pd
import urllib.request
import time
import random
import re
from keras.preprocessing.text import Tokenizer
from sklearn.model_selection import train_test_split
import numpy as np
import json
import codecs









def write_one_row(data_list, filename):
    """
    write_one_row(['hash', 'vin_num', 'vout_num',  'in_address', 'out_address', 'type', 'value', 'fees', 'opreturn'], PATH_FAIL)
    :param data_list:
    :param filename:
    :return:
    """
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet1']
    treedata1 = [data_list]
    for row in range(len(treedata1)):
        ws.append(treedata1[row])
    wb.save(filename)
    print(f'FILE OUT.xlsx===============================================================has insert {data_list}')

def Z_ScoreNormalization(x,mu,sigma):
    x = (x - mu) / sigma;
    return x;

def get_data():
    # filename_1 = 'vdata/train_features.xlsx'
    # filename_2 = 'vdata/train_rawdata.xlsx'
    # filename_3 = 'vdata/test_features.xlsx'
    # filename_4 = 'vdata/test_rawdata.xlsx'
    filename_2 = "gan/train.xlsx"
    filename_4 = "gan/test.xlsx"

    normal_data = []
    gen_data =[]
    special_data = []
    json_num = 0
    #处理json文件

    special_rawdata = "gan/gen_225.json"
    with open(special_rawdata, 'r', encoding='utf8')as fn:
        for line in fn.readlines():
            special = json.dumps(line)
            special = special.replace('\\n', '')
            special = special.replace('\\', '')
            special = special.replace('\"', '')
            special = special.replace('{', '')
            special = special.replace('}', '')
            special = special.replace('[', '')
            special = special.replace(']', '')
            special = special.replace(' ', '')
            special = special.replace(':', '')
            special = " ".join(special)
            special_data.append(special)

    normal_rawdata = "gan/1_2_225.json"
    with open(normal_rawdata, 'r', encoding='utf8')as fn:
        for line in fn.readlines():
            normal = json.dumps(line)
            normal = normal.replace('\\n', '')
            normal = normal.replace('\\', '')
            normal = normal.replace('\"', '')
            normal = normal.replace('{', '')
            normal = normal.replace('}', '')
            normal = normal.replace('[', '')
            normal = normal.replace(']', '')
            normal = normal.replace(' ', '')
            # normal = normal.replace('outputs', '')
            normal = normal.replace(':', '')
            normal = " ".join(normal)
            normal_data.append(normal)

    # for data in normal_data:
    #     with open("trian_gan.txt", "a", encoding='utf-8') as f1:
    #         f1.write(data)
    #         f1.write("\n")

    # with open("gan/gen_12.txt",'r',encoding='utf8') as fg:
    #     for line in fg.readlines():
    #
    #         gen_data.append(line)





    print(len(normal_data))
    print(len(special_data))

    # #处理特征值数据，将数据写入list
    # normal_feature = "vdata/btc_normal_features.xlsx"
    # normal = openpyxl.load_workbook(normal_feature)
    # normal = normal.active
    # no_feat = list()
    # no_rows = normal.max_row
    # i = 0
    # while (i < no_rows):
    #     row = []
    #     for cell in list(normal.rows)[i]:  # 遍历数据行
    #         row.append(cell.value)
    #     no_feat.append([row[0], row[1], row[2], row[3], row[4], row[5]])
    #     i += 1
    # print(len(no_feat))
    #
    # normal_feature = "vdata/647281_feature.xlsx"
    # normal = openpyxl.load_workbook(normal_feature)
    # normal = normal.active
    # no_rows = normal.max_row
    # i = 0
    # while (i < no_rows):
    #     row = []
    #     for cell in list(normal.rows)[i]:  # 遍历数据行
    #         row.append(cell.value)
    #     no_feat.append([row[0], row[1], row[2], row[3], row[4], row[5]])
    #     i += 1
    # print(len(no_feat))
    #
    # xlsx_num = 0
    # sp_feat = list()
    # while(xlsx_num < 7):
    #     special_feature = "vdata/"+str(xlsx_num+1)+".xlsx"
    #     print(special_feature)
    #     special = openpyxl.load_workbook(special_feature)
    #     special= special.active
    #     sp_rows = special.max_row
    #
    #     i = 0
    #     while ( i < sp_rows):
    #         row = []
    #         for cell in list(special.rows)[i]:  # 遍历数据行
    #             row.append(cell.value)
    #         sp_feat.append([row[0], row[1], row[2], row[3], row[4], row[5]])
    #         i += 1
    #     xlsx_num += 1
    # print(len(sp_feat))






    # ##CNN的训练和测试数据


    train_2 = openpyxl.load_workbook(filename_2)
    train_rawdata = train_2.active
    test_2 = openpyxl.load_workbook(filename_4)
    test_rawdata = test_2.active

    no = 0
    sp = 0
    num = 1 #记录表格的行数
    count = 0
    #
    #训练数据
    while count < 420:



        print("train_normal")
        i = random.randint(1, 2)
        while i > 0 and no < 210:
            string = normal_data[no]
            # string_1 = " ".join(string)
            train_rawdata.cell(row=num, column=1, value=string)
            train_rawdata.cell(row=num, column=2, value=int(0))
            # write_one_row(no_feat[no], filename_1)

            num = num + 1
            i = i - 1
            no = no + 1
            count = count + 1
            print(num)

        print("trian_special")
        j = random.randint(1, 2)
        while j > 0 and sp < 210:

            # train_rawdata.cell(row=num, column=1, value=gen_data[sp])
            # train_rawdata.cell(row=num, column=2, value=int(1))
            train_rawdata.cell(row=num, column=1, value=special_data[sp])
            train_rawdata.cell(row=num, column=2, value=int(1))

            num = num + 1
            j = j - 1
            sp = sp + 1
            count = count + 1
            print(num)

    # #
    # #
    # #
    # # #测试数据
    num = 1
    count = 0
    while count < 180:
        print("test_normal")
        i = random.randint(1, 2)
        while i > 0 and no < 500:
            string = normal_data[no]
            # string_1 = " ".join(string)
            test_rawdata.cell(row=num, column=1, value=string)
            test_rawdata.cell(row=num, column=2, value=int(0))
            # write_one_row(no_feat[no], filename_3)

            num = num + 1
            i = i - 1
            no = no + 1
            count = count + 1
        print("trian_special")
        j = random.randint(1, 2)
        while j > 0 and sp < 500:
            # test_rawdata.cell(row=num, column=1, value=gen_data[sp])
            # test_rawdata.cell(row=num, column=2, value=int(1))
            test_rawdata.cell(row=num, column=1, value=special_data[sp])
            test_rawdata.cell(row=num, column=2, value=int(1))
            num = num + 1
            j = j - 1
            sp = sp + 1
            count = count + 1
            print(num)






    train_2.save(filename_2)
    test_2.save(filename_4)



def load_data():
    filename_1 = 'vdata/train_features.xlsx'
    filename_2 = 'vdata/train_rawdata.xlsx'
    filename_3 = 'vdata/test_features.xlsx'
    filename_4 = 'vdata/test_rawdata.xlsx'
    train_data = []
    train_label =[]
    test_data =[]
    test_label =[]
    train = openpyxl.load_workbook(filename_2)
    train = train.active
    test= openpyxl.load_workbook(filename_4)
    test = test.active
    for cell in list(train.columns)[0]:  #获取第三列的数据
        train_data.append(cell.value)

    for cell in list(train.columns)[1]:  #获取第三列的数据
        train_label.append(cell.value)

    for cell in list(test.columns)[0]:  #获取第三列的数据
        test_data.append(cell.value)
    for cell in list(test.columns)[1]:  #获取第三列的数据
        test_label.append(cell.value)

    data = train_data
    data.extend(test_data)
    print(train_label)
    # tokenizer = Tokenizer(num_words=1000,filters='!"#$%&()*+,-./:;<=>?@[\]^_`{|}~\t\n',char_level=True)  # i创建一个分词器（tokenizer），设置为只考虑前1000个最常见的单词
    tokenizer = Tokenizer(num_words=1000)
    tokenizer.fit_on_texts(data)   # 构建索引单词

    sequences = tokenizer.texts_to_sequences(data)   # 将字符串转换为整数索引组成的列表
    data_all = []
    for string in sequences :
        data_all.append(string)

    train = data_all[0:5250]
    test = data_all[5250:7500]

    train = np.array(train)
    test = np.array(test)
    train_label = np.array(train_label)
    test_label = np.array(test_label)



    train_fea = openpyxl.load_workbook(filename_1)
    test_fea = openpyxl.load_workbook(filename_3)
    train_fea = train_fea.active
    test_fea = test_fea.active

    train_input = [[0.0 for i in range(5250)] for i in range(6)]
    test_input = [[0.0 for i in range(2250)] for i in range(6)]

    train_input = np.array(train_input)
    test_input = np.array(test_input)
    # 获取列数据
    i = 0
    for col in train_fea.columns:
        col_val = [row.value for row in col]
        col_np = np.array(col_val)
        if (i < 6):
            train_input[[i], :] = col_np
        i += 1

    i = 0
    for col in test_fea.columns:
        col_val = [row.value for row in col]
        col_np = np.array(col_val)
        if (i < 6):
            test_input[[i], :] = col_np
        i += 1

    i  = 0
    while i < len(train_input):
        train_input[i] =Z_ScoreNormalization(train_input[i],train_input[i].mean(),train_input[i].std())
        i += 1
    while i < len(test_input):
        test_input[i] =Z_ScoreNormalization(test_input[i],test_input[i].mean(),test_input[i].std())
        i += 1


    return(train,train_input.T,train_label,test,test_input.T,test_label)

def load_BP_data():
    filename_1 = 'vdata/train_features.xlsx'

    filename_3 = 'vdata/test_features.xlsx'

    train_data = []
    train_label = []
    test_data = []
    test_label = []
    train_fea = openpyxl.load_workbook(filename_1)
    test_fea = openpyxl.load_workbook(filename_3)
    train_fea = train_fea.active
    test_fea = test_fea.active

    train_input = [[0.0 for i in range(7140)] for i in range(6)]
    test_input = [[0.0 for i in range(3060)] for i in range(6)]

    train_input = np.array(train_input)
    test_input = np.array(test_input)

    for cell in list(train_fea.columns)[6]:  #获取第三列的数据
        train_label.append(cell.value)


    for cell in list(test_fea.columns)[6]:  #获取第三列的数据
        test_label.append(cell.value)
    # 获取列数据
    i = 0
    for col in train_fea.columns:
        col_val = [row.value for row in col]
        col_np = np.array(col_val)
        if (i < 6):
            train_input[[i], :] = col_np
        i += 1

    i = 0
    for col in test_fea.columns:
        col_val = [row.value for row in col]
        col_np = np.array(col_val)
        if (i < 6):
            test_input[[i], :] = col_np
        i += 1

    i = 0
    while i < len(train_input):
        train_input[i] = Z_ScoreNormalization(train_input[i], train_input[i].mean(), train_input[i].std())
        i += 1
    while i < len(test_input):
        test_input[i] = Z_ScoreNormalization(test_input[i], test_input[i].mean(), test_input[i].std())
        i += 1

    return ( train_input.T, train_label,  test_input.T, test_label)
def load_CNN_data():

    filename_2 = '10.7data/4_train_value.xlsx'

    filename_4 = '10.7data/4_test_value.xlsx'
    train_data = []
    train_label = []
    test_data = []
    test_label = []
    train = openpyxl.load_workbook(filename_2)
    train = train.active
    test = openpyxl.load_workbook(filename_4)
    test = test.active
    for cell in list(train.columns)[0]:  # 获取第三列的数据
        train_data.append(cell.value)

    for cell in list(train.columns)[1]:  # 获取第三列的数据
        train_label.append(cell.value)

    for cell in list(test.columns)[0]:  # 获取第三列的数据
        test_data.append(cell.value)
    for cell in list(test.columns)[1]:  # 获取第三列的数据
        test_label.append(cell.value)

    data = train_data
    data.extend(test_data)
    print(train_label)
    # tokenizer = Tokenizer(num_words=1000,filters='!"#$%&()*+,-./:;<=>?@[\]^_`{|}~\t\n',char_level=True)  # i创建一个分词器（tokenizer），设置为只考虑前1000个最常见的单词
    tokenizer = Tokenizer(num_words=1000)
    tokenizer.fit_on_texts(data)  # 构建索引单词

    sequences = tokenizer.texts_to_sequences(data)  # 将字符串转换为整数索引组成的列表
    data_all = []
    for string in sequences:
        data_all.append(string)

    train = data_all[0:1400]
    test = data_all[1400:2000]

    return (train, train_label, test,  test_label)

